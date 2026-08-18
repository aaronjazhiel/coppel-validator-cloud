"""Worker de validación N2 — Lee el .xlsx de S3 y valida completitud por hoja."""
import json
import re
from io import BytesIO
from urllib.parse import unquote_plus

from openpyxl import load_workbook
from common.helpers import DynamoHelper, S3Helper, log_event


# ── Reglas de validación por hoja ─────────────────────────────
REGLAS_REQ_INFRA = {
    "nombre": "Req Infra-Ambiente",
    "campos_obligatorios": [
        ("Nombre del proyecto*", 6, 3),
        ("Ambiente*", 8, 3),
        ("Categoría MMA Asignada*", 12, 3),
        ("RTO*", 13, 3),
        ("RPO*", 14, 3),
        ("Analisis de Vulneravilidades*", 15, 3),
        ("Documento Continuidad TI*", 16, 3),
        ("La aplicacion puede trabajar con HA*", 17, 3),
        ("La App Trabaja con DNS*", 18, 3),
        ("Trabaja con datos Encriptados*", 19, 3),
        ("Plan de Continuidad*", 20, 3),
        ("Nube pública*", 25, 3),
    ]
}

# Secciones de recursos AWS a validar
SECCIONES_AWS = [
    {"nombre": "EC2", "hoja": "(N2,D) Recursos AWS", "fila_inicio": 1, "campos": ["Sistema Operativo", "Vcpu", "Memory", "Storage"]},
    {"nombre": "RDS", "hoja": "(N2,D) Recursos AWS", "fila_inicio": 17, "campos": ["Nombre", "Motor BD", "Version", "Vcpu", "Memory", "Storage"]},
    {"nombre": "EKS", "hoja": "(N2,D) Recursos AWS", "fila_inicio": 30, "campos": ["Version de EKS", "Numero de nodos"]},
]

SECCIONES_GCP = [
    {"nombre": "Instancias VM", "hoja": "(N2,D) Recursos GCP"},
    {"nombre": "Cloud SQL", "hoja": "Cloud SQL DB"},
    {"nombre": "GKE", "hoja": "GKE "},
    {"nombre": "Pub/Sub", "hoja": "Pub-Sub"},
    {"nombre": "Mongo Atlas", "hoja": "Mongo Atlas"},
    {"nombre": "AlloyDB", "hoja": "AlloyDB"},
    {"nombre": "Dataflow", "hoja": "Datafow"},
    {"nombre": "Composer", "hoja": "Composer"},
    {"nombre": "Firestore", "hoja": "Firestore"},
    {"nombre": "BigQuery", "hoja": "BigQuery "},
    {"nombre": "Cloud Storage", "hoja": "Cloud Storage"},
    {"nombre": "DataProc", "hoja": "DataProc"},
]

VALORES_VACIOS = {"", "Seleccione", "seleccione", "None", "none", "Link Documento"}


def validar_n2(id_iniciativa: str, meta: dict):
    """Valida el Excel N2 desde S3 y guarda resultados."""
    insumos = meta.get("insumos") or []
    n2_keys = [unquote_plus(i["key"]) for i in insumos if i.get("tipo") == "componentes"]

    if not n2_keys:
        resultado = {
            "puntaje_total": 0,
            "estado": "SIN_ARCHIVO",
            "mensaje": "No se subió ningún archivo de Componentes N2",
            "secciones": {}
        }
        S3Helper.put_json(f"iniciativas/{id_iniciativa}/resultados/validacion_n2.json", resultado)
        DynamoHelper.put_event(id_iniciativa, "N2_SIN_ARCHIVO")
        return resultado

    # Leer el primer .xlsx
    raw = S3Helper.read_bytes(n2_keys[0])
    wb = load_workbook(BytesIO(raw), data_only=True)

    resultado = {
        "puntaje_total": 0,
        "secciones": {},
        "recursos_detectados": {},
        "hojas_disponibles": wb.sheetnames,
    }

    # 1. Validar Req Infra-Ambiente
    resultado["secciones"]["req_infra_ambiente"] = _validar_req_infra(wb)

    # 2. Validar Recursos AWS
    resultado["secciones"]["recursos_aws"] = _validar_recursos_aws(wb)

    # 3. Validar Recursos GCP
    resultado["secciones"]["recursos_gcp"] = _validar_recursos_gcp(wb)

    # 4. Detectar recursos usados
    resultado["recursos_detectados"] = _detectar_recursos(wb)

    # Calcular puntaje total ponderado
    pesos = {"req_infra_ambiente": 40, "recursos_aws": 30, "recursos_gcp": 30}
    total = 0
    for sec, peso in pesos.items():
        p = resultado["secciones"].get(sec, {}).get("puntaje", 0)
        total += p * (peso / 100)
    resultado["puntaje_total"] = round(total)

    # Guardar
    S3Helper.put_json(f"iniciativas/{id_iniciativa}/resultados/validacion_n2.json", resultado)
    DynamoHelper.put_event(id_iniciativa, "N2_VALIDADO", {"puntaje": resultado["puntaje_total"]})
    log_event("validar_n2", {"id_iniciativa": id_iniciativa, "puntaje": resultado["puntaje_total"]})

    return resultado


def _validar_req_infra(wb):
    """Valida campos obligatorios de la hoja Req Infra-Ambiente."""
    sheet_name = "(N2,D) Req Infra-Ambiente"
    if sheet_name not in wb.sheetnames:
        return {"puntaje": 0, "estado": "HOJA_NO_ENCONTRADA", "campos": []}

    ws = wb[sheet_name]
    campos = []
    completos = 0

    for nombre, fila, col in REGLAS_REQ_INFRA["campos_obligatorios"]:
        valor = ws.cell(row=fila, column=col).value
        valor_str = str(valor).strip() if valor else ""
        es_valido = valor_str not in VALORES_VACIOS

        campos.append({
            "campo": nombre,
            "estado": "COMPLETO" if es_valido else "VACIO",
            "valor": valor_str if es_valido else "",
        })
        if es_valido:
            completos += 1

    total = len(campos)
    puntaje = round((completos / total) * 100) if total else 0

    return {"puntaje": puntaje, "campos": campos, "completos": completos, "total": total}


def _validar_recursos_aws(wb):
    """Valida la hoja de Recursos AWS — EC2, RDS, EKS."""
    sheet_name = "(N2,D) Recursos AWS"
    if sheet_name not in wb.sheetnames:
        return {"puntaje": 0, "estado": "HOJA_NO_ENCONTRADA", "subsecciones": []}

    ws = wb[sheet_name]
    subsecciones = []

    # EC2: filas 5-7 (datos), columnas B-G
    ec2_items = _extraer_tabla(ws, fila_inicio=5, fila_fin=10, col_inicio=2, col_fin=7,
                               headers=["SO", "Distribucion", "Version", "Vcpu", "Memory", "Storage"])
    ec2_validos = [r for r in ec2_items if any(str(v).strip() not in VALORES_VACIOS for v in r.values() if v)]
    subsecciones.append({
        "nombre": "EC2",
        "instancias": len(ec2_validos),
        "estado": "CONFIGURADO" if ec2_validos else "VACIO",
        "detalle": ec2_validos
    })

    # RDS: buscar sección
    rds_items = _extraer_tabla(ws, fila_inicio=21, fila_fin=25, col_inicio=2, col_fin=7,
                               headers=["Nombre", "Motor BD", "Version", "Vcpu", "Memory", "Storage"])
    rds_validos = [r for r in rds_items if any(str(v).strip() not in VALORES_VACIOS for v in r.values() if v)]
    subsecciones.append({
        "nombre": "RDS",
        "instancias": len(rds_validos),
        "estado": "CONFIGURADO" if rds_validos else "VACIO",
        "detalle": rds_validos
    })

    # EKS: buscar sección
    eks_data = {}
    for row in range(30, min(50, ws.max_row)):
        cell_b = ws.cell(row=row, column=2).value
        if cell_b and "Version de EKS" in str(cell_b):
            eks_data["version"] = str(ws.cell(row=row, column=3).value or "")
        if cell_b and "Numero de nodos" in str(cell_b):
            eks_data["nodos"] = str(ws.cell(row=row, column=3).value or "")

    subsecciones.append({
        "nombre": "EKS",
        "estado": "CONFIGURADO" if any(eks_data.values()) else "VACIO",
        "detalle": eks_data
    })

    # Puntaje: % de subsecciones con datos
    configurados = sum(1 for s in subsecciones if s["estado"] == "CONFIGURADO")
    # Si al menos una tiene datos, dar puntaje proporcional
    puntaje = round((configurados / len(subsecciones)) * 100) if subsecciones else 0

    return {"puntaje": puntaje, "subsecciones": subsecciones}


def _validar_recursos_gcp(wb):
    """Valida hojas de recursos GCP."""
    subsecciones = []

    for sec in SECCIONES_GCP:
        hoja = sec["hoja"]
        if hoja not in wb.sheetnames:
            subsecciones.append({"nombre": sec["nombre"], "estado": "HOJA_NO_ENCONTRADA"})
            continue

        ws = wb[hoja]
        # Contar filas con datos reales (ignorar headers y vacías)
        filas_con_datos = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            valores = [str(c).strip() for c in row if c and str(c).strip() not in VALORES_VACIOS]
            if len(valores) >= 2:
                filas_con_datos += 1

        subsecciones.append({
            "nombre": sec["nombre"],
            "estado": "CONFIGURADO" if filas_con_datos > 0 else "VACIO",
            "filas_datos": filas_con_datos
        })

    configurados = sum(1 for s in subsecciones if s["estado"] == "CONFIGURADO")
    # Puntaje basado en si hay al menos algún recurso GCP configurado
    puntaje = round((configurados / max(len(subsecciones), 1)) * 100) if configurados > 0 else 0

    return {"puntaje": puntaje, "subsecciones": subsecciones}


def _detectar_recursos(wb):
    """Resumen de qué recursos están configurados."""
    recursos = {"aws": [], "gcp": []}

    if "(N2,D) Recursos AWS" in wb.sheetnames:
        ws = wb["(N2,D) Recursos AWS"]
        # Buscar secciones con datos
        for row in range(1, min(ws.max_row, 100)):
            val = ws.cell(row=row, column=2).value
            if val and any(kw in str(val) for kw in ["EC2", "RDS", "EKS", "S3", "Lambda", "ElastiCache"]):
                recursos["aws"].append(str(val).strip())

    for sec in SECCIONES_GCP:
        if sec["hoja"] in wb.sheetnames:
            ws = wb[sec["hoja"]]
            filas = sum(1 for row in ws.iter_rows(min_row=5, values_only=True)
                       if any(c and str(c).strip() not in VALORES_VACIOS for c in row))
            if filas > 0:
                recursos["gcp"].append(sec["nombre"])

    return recursos


def _extraer_tabla(ws, fila_inicio, fila_fin, col_inicio, col_fin, headers):
    """Extrae filas de una tabla del Excel."""
    items = []
    for row in range(fila_inicio, fila_fin + 1):
        valores = {}
        tiene_dato = False
        for i, col in enumerate(range(col_inicio, col_fin + 1)):
            val = ws.cell(row=row, column=col).value
            val_str = str(val).strip() if val else ""
            if i < len(headers):
                valores[headers[i]] = val_str
            if val_str and val_str not in VALORES_VACIOS:
                tiene_dato = True
        if tiene_dato:
            items.append(valores)
    return items
