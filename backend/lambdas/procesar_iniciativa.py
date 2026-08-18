"""POST /iniciativas/{id}/procesar — Extracción asíncrona con Claude.

Lee todos los insumos de una iniciativa desde S3, los convierte a texto
y envía a Claude para extraer la Ficha Técnica estructurada.
Usa auto-invocación async para no bloquear el API Gateway (timeout 29s).
"""
import json
import time
import io
from xml.etree import ElementTree as ET

import boto3
from docx import Document
from openpyxl import load_workbook

from common.schema import FichaTecnica
from common.claude_client import ClaudeClient
from common.helpers import build_response, S3Helper, DynamoHelper, log_event

lambda_client = boto3.client("lambda")


def handler(event, context):
    """Entry point: responde 202 al API Gateway e invoca a sí misma de forma async."""
    # Si es invocación directa asíncrona, ejecutar el procesamiento
    if event.get("_async_process"):
        return _process(event["id_iniciativa"])

    # Desde API Gateway
    id_iniciativa = event.get("pathParameters", {}).get("id", "")
    if not id_iniciativa:
        return build_response(400, {"error": "id_iniciativa requerido"})

    meta = DynamoHelper.get_meta(id_iniciativa)
    if not meta:
        return build_response(404, {"error": "Iniciativa no encontrada"})

    # Auto-invocación async para procesamiento largo sin timeout de API Gateway
    lambda_client.invoke(
        FunctionName=context.function_name,
        InvocationType="Event",
        Payload=json.dumps({"_async_process": True, "id_iniciativa": id_iniciativa}),
    )

    DynamoHelper.update_meta(id_iniciativa, {"estado": "PROCESANDO", "GSI1PK": "ESTADO#PROCESANDO", "fecha_actualizacion": DynamoHelper.now_iso()})
    return build_response(202, {"id_iniciativa": id_iniciativa, "status": "PROCESANDO"})


def _process(id_iniciativa: str):
    """Lee insumos de S3, los convierte a texto y extrae la Ficha Técnica con Claude."""
    start = time.time()
    try:
        meta = DynamoHelper.get_meta(id_iniciativa)
        insumos = meta.get("insumos", [])
        prefix = meta.get("s3_prefix", f"iniciativas/{id_iniciativa}")

        # Leer todos los insumos
        contenido_partes = []
        for insumo in insumos:
            key = insumo["key"]
            tipo = insumo["tipo"]
            try:
                text = _read_insumo(key, tipo)
                if text:
                    contenido_partes.append(f"=== FUENTE: {tipo} ({key.split('/')[-1]}) ===\n{text}")
            except Exception as e:
                log_event("error_lectura_insumo", {"key": key, "error": str(e)})

        if not contenido_partes:
            # Intentar listar archivos directamente de S3
            for tipo in ["discovery", "diagrama", "transcripciones", "componentes"]:
                keys = S3Helper.list_keys(f"{prefix}/insumos/{tipo}/")
                for key in keys:
                    try:
                        text = _read_insumo(key, tipo)
                        if text:
                            contenido_partes.append(f"=== FUENTE: {tipo} ({key.split('/')[-1]}) ===\n{text}")
                    except Exception:
                        pass

        if not contenido_partes:
            DynamoHelper.update_meta(id_iniciativa, {"estado": "ERROR", "GSI1PK": "ESTADO#ERROR", "fecha_actualizacion": DynamoHelper.now_iso()})
            DynamoHelper.put_event(id_iniciativa, "ERROR", {"detalle": "Sin insumos para procesar"})
            return

        prompt = (
            "A continuación se presentan los insumos de un proyecto de arquitectura cloud. "
            "EXTRAE la información y estructura la Ficha Técnica. "
            "CONCILIA las fuentes: si un dato aparece en varias fuentes, consolida. "
            "Si hay discrepancia, regístralo como conflicto. "
            "NO inventes datos: lo ausente debe quedar vacío y listado en validacion.huecos.\n\n"
            + "\n\n".join(contenido_partes)
        )

        # Obtener schema JSON del modelo Pydantic
        schema = FichaTecnica.model_json_schema()

        client = ClaudeClient()
        result = client.extract_ficha(prompt, schema)
        ficha_data = result["ficha"]
        usage = result["usage"]

        # Validar con Pydantic
        ficha = FichaTecnica.model_validate(ficha_data)
        ficha_dict = ficha.model_dump()

        # Guardar ficha.json
        S3Helper.put_json(f"{prefix}/ficha/ficha.json", ficha_dict)

        duracion_ms = int((time.time() - start) * 1000)

        DynamoHelper.update_meta(id_iniciativa, {
            "estado": "EXTRAIDO",
            "GSI1PK": "ESTADO#EXTRAIDO",
            "GSI1SK": DynamoHelper.now_iso(),
            "completitud": ficha.validacion.completitud,
            "huecos": ficha.validacion.huecos,
            "fecha_actualizacion": DynamoHelper.now_iso(),
        })
        DynamoHelper.put_event(id_iniciativa, "PROCESADO", {
            "modelo": client.model,
            "tokens_in": usage.get("input_tokens", 0),
            "tokens_out": usage.get("output_tokens", 0),
            "duracion_ms": duracion_ms,
        })
        log_event("procesar_ok", {"id_iniciativa": id_iniciativa, "duracion_ms": duracion_ms})

    except Exception as e:
        DynamoHelper.update_meta(id_iniciativa, {"estado": "ERROR", "GSI1PK": "ESTADO#ERROR", "fecha_actualizacion": DynamoHelper.now_iso()})
        DynamoHelper.put_event(id_iniciativa, "ERROR", {"detalle": str(e)})
        log_event("procesar_error", {"id_iniciativa": id_iniciativa, "error": str(e)})


def _read_insumo(key: str, tipo: str) -> str:
    """Lee un insumo de S3 y lo convierte a texto."""
    if tipo == "discovery":
        data = S3Helper.read_bytes(key)
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    elif tipo == "componentes":
        data = S3Helper.read_bytes(key)
        wb = load_workbook(io.BytesIO(data), read_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"[Hoja: {sheet}]")
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join(str(c) if c is not None else "" for c in row))
        return "\n".join(lines)

    elif tipo == "diagrama":
        xml_text = S3Helper.read_text(key)
        # Extraer labels de nodos y aristas del XML drawio
        try:
            root = ET.fromstring(xml_text)
            cells = root.iter("mxCell")
            labels = [c.get("value", "") for c in cells if c.get("value")]
            return f"[Diagrama draw.io — elementos:]\n" + "\n".join(labels)
        except ET.ParseError:
            return xml_text[:5000]

    else:  # transcripciones / texto plano
        return S3Helper.read_text(key)
